#----------------------------------------------------------------------------#

# Purpose:     Parse & fill out background
# Project:     NLP Tool/Christmas Present
# Author:      Clara Marquardt
# Date:        2016

#----------------------------------------------------------------------------#

#----------------------------------------------------------------------------#
#                               Control Section                              #
#----------------------------------------------------------------------------#

# dependencies
#-------------------------------------------------#
import pdfminer

from pdfminer.pdfparser import PDFParser
from pdfminer.pdfdocument import PDFDocument
from pdfminer.pdfpage import PDFPage
from pdfminer.pdfpage import PDFTextExtractionNotAllowed
from pdfminer.pdfinterp import PDFResourceManager
from pdfminer.pdfinterp import PDFPageInterpreter
from pdfminer.pdfdevice import PDFDevice
from pdfminer.layout import LAParams
from pdfminer.converter import PDFPageAggregator

from openpyxl import load_workbook
from openpyxl import Workbook

import os
import sys

import re

import glob

from random import randint


from PyPDF2 import PdfFileWriter, PdfFileReader

import StringIO

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.graphics.shapes import Rect
from reportlab.pdfgen.canvas import Canvas
from reportlab.lib.colors import PCMYKColor, PCMYKColorSep, Color, black, blue, red, white, green

import numpy as np
import pandas as pd
import math 

# control parameters
#-------------------------------------------------#
print 'Number of arguments:', len(sys.argv), 'arguments.'
# print 'Argument List:', str(sys.argv)

# paths
input_path=sys.argv[1]
print input_path

# paths
mod_path=sys.argv[2]
print mod_path

output_path=sys.argv[3]
print output_path

doc_path=sys.argv[4]
print doc_path

doc_path=doc_path+"/output/*.xlsm"

#----------------------------------------------------------------------------#
#                          Methods/Functions                                 #
#----------------------------------------------------------------------------#

# pdfPositionHandling - parse & extract line positions
#----------------------------------------------------------------------------#
class pdfPositionHandling:

    def parse_obj(self, lt_objs, pos_list, page_num):
    	position_list_new=pos_list

        # loop over the object list
        for obj in lt_objs:

            if isinstance(obj, pdfminer.layout.LTTextLine):
            	# print(obj.get_text().replace('\n', '_'))
                # print "%6d, %6d, %s" % (obj.bbox[0], obj.bbox[1], obj.get_text().replace('\n', '_'))
				# append to list
             	position_list_new.append([obj.bbox[0], obj.bbox[1], page_num,obj.get_text().replace('\n', '')])

            # if it's a textbox, also recurse
            if isinstance(obj, pdfminer.layout.LTTextBoxHorizontal):
                self.parse_obj(obj._objs, position_list_new, page_num)

            # if it's a container, recurse
            elif isinstance(obj, pdfminer.layout.LTFigure):
                self.parse_obj(obj._objs,position_list_new, page_num)

    def parsepdf(self, filename, startpage, endpage):

        # Open a PDF file.
        fp = open(filename, 'rb')

        # Create Position List
        position_list=[]

        # Create a PDF parser object associated with the file object.
        parser = PDFParser(fp)

        # Create a PDF document object that stores the document structure.
        # Password for initialization as 2nd parameter
        document = PDFDocument(parser)

        # Check if the document allows text extraction. If not, abort.
        if not document.is_extractable:
            raise PDFTextExtractionNotAllowed

        # Create a PDF resource manager object that stores shared resources.
        rsrcmgr = PDFResourceManager()

        # Create a PDF device object.
        device = PDFDevice(rsrcmgr)

        # BEGIN LAYOUT ANALYSIS
        # Set parameters for analysis.
        laparams = LAParams()

        # Create a PDF page aggregator object.
        device = PDFPageAggregator(rsrcmgr, laparams=laparams)

        # Create a PDF interpreter object.
        interpreter = PDFPageInterpreter(rsrcmgr, device)


        i = 0
        # loop over all pages in the document
        for page in PDFPage.create_pages(document):
            if i >= startpage and i <= endpage:
                # read the page into a layout object
                interpreter.process_page(page)
                layout = device.get_result()

                # extract text from this object
                # print(position_list)
                self.parse_obj(layout._objs,position_list, i)
                i += 1

        position_list = pd.DataFrame(position_list)
        position_list.columns=["pos_x", "pos_y", "page", "text"]

        return(position_list)


#----------------------------------------------------------------------------#
#                                    Code                                    #
#----------------------------------------------------------------------------#


# read in xlsx 
#----------------------------------------------------------------------------#
output_file_list = glob.glob(doc_path)

order_name=[]
order_product_code_1=[]
order_product_code_2=[]
order_product_code_3=[]
order_price_1=[]
order_price_2=[]
order_price_3=[]
order_id=[]
email=[]

for x in range(0, len(output_file_list)):

    # read in 
    temp=load_workbook(output_file_list[x])
    temp_ws=temp['MASTER RECORD']

    # obtain relevant values
    temp_order_name=[temp_ws.cell(row=i,column=9).value for i in range(1,temp_ws.max_row)]
    temp_order_product_code_1=[temp_ws.cell(row=i,column=3).value for i in range(1,temp_ws.max_row)]
    temp_order_product_code_2=[temp_ws.cell(row=i,column=5).value for i in range(1,temp_ws.max_row)]
    temp_order_product_code_3=[temp_ws.cell(row=i,column=7).value for i in range(1,temp_ws.max_row)]
    temp_order_price_1=[temp_ws.cell(row=i,column=4).value for i in range(1,temp_ws.max_row)]
    temp_order_price_2=[temp_ws.cell(row=i,column=6).value for i in range(1,temp_ws.max_row)]
    temp_order_price_3=[temp_ws.cell(row=i,column=8).value for i in range(1,temp_ws.max_row)]

    temp_order_id=[temp_ws.cell(row=i,column=10).value for i in range(1,temp_ws.max_row)]

    temp_email=[temp_ws.cell(row=i,column=13).value for i in range(1,temp_ws.max_row)]

    # append
    order_name.append(temp_order_name)
    order_product_code_1.append(temp_order_product_code_1)
    order_product_code_2.append(temp_order_product_code_2)
    order_product_code_3.append(temp_order_product_code_3)
    order_price_1.append(temp_order_price_1)
    order_price_2.append(temp_order_price_2)
    order_price_3.append(temp_order_price_3)

    order_id.append(temp_order_id)
    
    email.append(temp_email)

## create final dt
order_dt=pd.DataFrame({'order_name': order_name[0][1:],
     'order_product_code_1': order_product_code_1[0][1:],
     'order_product_code_2': order_product_code_2[0][1:],
     'order_product_code_3': order_product_code_3[0][1:],
     'order_price_1': order_price_1[0][1:],
     'order_price_2': order_price_2[0][1:],
     'order_price_3': order_price_3[0][1:],
     'order_id': order_id[0][1:], 
     'email': email[0][1:]
    })


# loop over
#----------------------------------------------------------------------------#
file_list_final = order_dt['order_name']
file_list_final = file_list_final.unique()

for x in range(0, len(file_list_final)):


    file_name_mod=mod_path + "/" + file_list_final[x] + '.pdf'
    file_name_raw=input_path + "/" + file_list_final[x] + '.pdf'

    print file_name_mod 

    # obtain order_dt_subset
    #----------------------------------------------------------------------------#
    order_dt_subset=order_dt.ix[order_dt['order_name']==file_list_final[x]]
    order_dt_subset.reset_index(inplace=True, drop=True)
    

    # parse PDF 
    #----------------------------------------------------------------------------#
    # read existing PDF
    existing_pdf             = PdfFileReader(file(file_name_mod, "rb"))
    existing_pdf_page_number = existing_pdf.getNumPages() 

    # parse
    position_class   = pdfPositionHandling()
    position         = position_class.parsepdf(file_name_mod, 0, existing_pdf_page_number)
    position_sort    = position.sort_values(['page', 'pos_y'],  ascending=[1, 0])
    position_sort.reset_index(inplace=True, drop=True)
    
    # read existing PDF - RAW
    existing_pdf_raw  = PdfFileReader(file(file_name_raw, "rb"))

    
    # new PDF 
    #----------------------------------------------------------------------------#
    can_list    = []
    packet_list = []
    pdf_list    = []


    for i in range(0,existing_pdf_page_number):
        packet_temp = StringIO.StringIO()
        can_temp    = canvas.Canvas(packet_temp)
        packet_list.append(packet_temp)
        can_list.append(can_temp)

    # iterate through order_dt_subset
    #----------------------------------------------------------------------------#

    for y in range(0, len(order_dt_subset)):

        # obtain data
        id_1=order_dt_subset['order_product_code_1'][y]
        id_2=order_dt_subset['order_product_code_2'][y]
        id_3=order_dt_subset['order_product_code_3'][y]

        price_1=order_dt_subset['order_price_1'][y]
        price_2=order_dt_subset['order_price_2'][y]
        price_3=order_dt_subset['order_price_3'][y]

        text_1 = str(id_1) + " / " + str(price_1)
        text_1=     re.sub("None|nan|#N/A", "", text_1)
        text_1=     re.sub("^ / $", "", text_1)
        text_2 = str(id_2) + " / " + str(price_2)
        text_2=     re.sub("None|nan|#N/A", "", text_2)
        text_2=     re.sub("^ / $", "", text_2)
        text_3 = str(id_3) + " / " + str(price_3)
        text_3=     re.sub("None|nan|#N/A", "", text_3)
        text_3=     re.sub("^ / $", "", text_3)

        order_id=order_dt_subset['order_id'][y]

        # identify the correct position 
        position_prod_subset=position_sort.copy()
        position_prod_subset.ix[position_prod_subset["text"].str.contains("#order-item: " + str(order_id) + "([^0-9]|$)"), "min"]=0
        position_prod_subset.ix[position_prod_subset["text"].str.contains("#order-item: " + str(order_id+1) + "([^0-9]|$)"), "min"]=1
        position_prod_subset.fillna(method="ffill", inplace=True)
        position_prod_subset=position_prod_subset.ix[position_prod_subset["min"]==0]


        min_index=position_prod_subset.index.min()
        max_index=position_prod_subset.index.max()
        id_line = position_prod_subset[position_prod_subset["text"].str.contains("(Stck( |$))|(Stk( |$))|(St( |$))")].index
        
        if (id_line.shape[0]>0):
            position_line=np.array(id_line)[0]
            x1    = position_sort.ix[(position_line)]["pos_x"]+60
            y1    = position_sort.ix[(position_line)]["pos_y"]+1
            y_1   = y1-10
            y_2   = y1-20

        else:
            position_line=min_index
            x1    = max(position_prod_subset["pos_x"])+10
            y1    = position_sort.ix[(min_index)]["pos_y"] - 20
            y_1   = y1-10
            y_2   = y1-20

        page_num = position_sort.ix[(position_line)]["page"]   


        # print
        #----------------------------------------------------------------------------#

        # iterate & draw line/add text
        can_page = can_list[page_num]
        can_page.setFont('Helvetica', 10)
        can_page.setFillColor("blue")
        can_page.drawString(x1, y1, text_1)
        can_page.drawString(x1, y_1, text_2)
        can_page.drawString(x1, y_2, text_3)      
          
    # merge with background PDF
    #----------------------------------------------------------------------------#
    for i in range(0,existing_pdf_page_number):
        can_list[i].save()
        packet_list[i].seek(0)
        pdf_list.append(PdfFileReader(packet_list[i]))

    output = PdfFileWriter()

    for i in range(0,existing_pdf_page_number):
        page = existing_pdf_raw.getPage(i)
        if pdf_list[i].getNumPages()>0:
            page.mergePage(pdf_list[i] .getPage(0))
        output.addPage(page)

    # save 
    outputStream = file(output_path + "/"+  file_list_final[x] + '.pdf', "wb")
    output.write(outputStream)
    outputStream.close()

    # save email list
    #----------------------------------------------------------------------------#
email_dt=order_dt[['email','order_name']]
email_dt=email_dt.drop_duplicates()
email_dt['order_name_mod'] = [re.sub("^[0-9]*_", "",x) for x in  email_dt['order_name']]
email_dt['order_name_mod'] = [re.sub("_", " ",x) for x in  email_dt['order_name_mod']]

# save
email_dt.to_csv(output_path + "/" + "email_list.csv", encoding="utf8")


#----------------------------------------------------------------------------#
#                                    End                                     #
#----------------------------------------------------------------------------#

